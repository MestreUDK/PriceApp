# routes/routes_core.py
from flask import (
    Blueprint, render_template, request, redirect, url_for, send_from_directory, current_app, Response
)
from extensions import db
from models import Supermercado, Produto, Preco, User, Categoria  # MUDANÇA: Categoria
from flask_login import login_required 
import json 
import openpyxl
from io import BytesIO 

core_bp = Blueprint('core', __name__,
                    template_folder='../templates',
                    static_folder='../static') 

@core_bp.route('/sw.js') 
def service_worker():
    return send_from_directory(current_app.static_folder, 'sw.js')

@core_bp.route('/')
@login_required 
def index():
    ultimos_precos = Preco.query.order_by(Preco.data_cadastro.desc()).limit(5).all()
    return render_template('index.html', ultimos_precos=ultimos_precos)

@core_bp.route('/busca')
@login_required 
def busca():
    termo = request.args.get('q')
    
    if not termo:
        return redirect(url_for('core.index'))
    
    termo_busca = f"%{termo}%"
    
    produtos_encontrados = Produto.query.filter(Produto.nome.ilike(termo_busca)).order_by(Produto.nome).all()
    mercados_encontrados = Supermercado.query.filter(Supermercado.nome.ilike(termo_busca)).order_by(Supermercado.nome).all()
    # MUDANÇA: Busca em Categorias em vez de Marcas
    categorias_encontradas = Categoria.query.filter(Categoria.nome.ilike(termo_busca)).order_by(Categoria.nome).all()
    
    return render_template('busca.html', 
                           termo=termo, 
                           produtos=produtos_encontrados, 
                           mercados=mercados_encontrados,
                           categorias=categorias_encontradas) # MUDANÇA: Passa categorias

@core_bp.route('/backup/download')
@login_required
def download_backup():
    # Esta rota JSON é mantida para o Leitor Offline
    usuarios = User.query.all()
    produtos = Produto.query.all()
    mercados = Supermercado.query.all()
    precos = Preco.query.all()
    categorias = Categoria.query.all()  # MUDANÇA: Categorias

    usuarios_list = [
        {"id": u.id, "username": u.username, "role": u.role}
        for u in usuarios
    ]
    produtos_list = [
        {"id": p.id, "nome": p.nome, "criado_por_id": p.criado_por_id, "editado_por_id": p.editado_por_id}
        for p in produtos
    ]
    mercados_list = [
        {"id": m.id, "nome": m.nome, "endereço": m.endereço, "criado_por_id": m.criado_por_id, "editado_por_id": m.editado_por_id}
        for m in mercados
    ]
    precos_list = [
        {"id": pr.id, "produto_id": pr.produto_id, "supermercado_id": pr.supermercado_id, 
         "categoria_id": pr.categoria_id,  # MUDANÇA: categoria_id
         "valor": pr.valor, "data_cadastro": pr.data_cadastro.isoformat(), "criado_por_id": pr.criado_por_id,
         "e_promocao": pr.e_promocao,
         "data_expiracao": pr.data_expiracao.isoformat() if pr.data_expiracao else None
        }
        for pr in precos
    ]
    categorias_list = [ # MUDANÇA: Categorias
        {"id": cat.id, "nome": cat.nome, "criado_por_id": cat.criado_por_id, "editado_por_id": cat.editado_por_id}
        for cat in categorias
    ]

    backup_data = {
        "usuarios": usuarios_list,
        "produtos": produtos_list,
        "supermercados": mercados_list,
        "precos": precos_list,
        "categorias": categorias_list # MUDANÇA
    }

    json_data = json.dumps(backup_data, indent=2, ensure_ascii=False)
    
    return Response(
        json_data,
        mimetype="application/json",
        headers={"Content-Disposition": "attachment;filename=priceapp_backup.json"}
    )

@core_bp.route('/backup/download_excel')
@login_required
def download_excel_backup():
    # 1. Cria o Workbook (o arquivo Excel)
    wb = openpyxl.Workbook()
    
    # 2. Remove a planilha padrão
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 3. Processa Usuários
    ws_users = wb.create_sheet("Usuarios")
    ws_users.append(["id", "username", "role", "email", "telefone"])
    for u in User.query.all():
        ws_users.append([u.id, u.username, u.role, u.email, u.telefone])

    # 4. Processa Produtos
    ws_prods = wb.create_sheet("Produtos")
    ws_prods.append(["id", "nome", "criado_por_id", "editado_por_id"])
    for p in Produto.query.all():
        ws_prods.append([p.id, p.nome, p.criado_por_id, p.editado_por_id])

    # 5. Processa Supermercados
    ws_markets = wb.create_sheet("Supermercados")
    ws_markets.append(["id", "nome", "endereço", "criado_por_id", "editado_por_id"])
    for m in Supermercado.query.all():
        ws_markets.append([m.id, m.nome, m.endereço, m.criado_por_id, m.editado_por_id])

    # 6. Processa Categorias (MUDANÇA)
    ws_cats = wb.create_sheet("Categorias")
    ws_cats.append(["id", "nome", "criado_por_id", "editado_por_id"])
    for c in Categoria.query.all():
        ws_cats.append([c.id, c.nome, c.criado_por_id, c.editado_por_id])

    # 7. Processa Preços
    ws_prices = wb.create_sheet("Precos")
    ws_prices.append([
        "id", "produto_id", "supermercado_id", "categoria_id", "valor",  # MUDANÇA: categoria_id
        "data_cadastro", "criado_por_id", "e_promocao", "data_expiracao"
    ])
    for pr in Preco.query.all():
        ws_prices.append([
            pr.id, pr.produto_id, pr.supermercado_id, pr.categoria_id, pr.valor,
            pr.data_cadastro, pr.criado_por_id, pr.e_promocao, pr.data_expiracao
        ])
    
    # 8. Salva em memória
    excel_file_memory = BytesIO()
    wb.save(excel_file_memory)
    excel_file_memory.seek(0)
    
    return Response(
        excel_file_memory.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=priceapp_backup.xlsx"}
    )

@core_bp.route('/leitor-offline')
@login_required 
def leitor_offline():
    return render_template('leitor_offline.html')

@core_bp.route('/health')
def health_check():
    return "App is awake", 200
